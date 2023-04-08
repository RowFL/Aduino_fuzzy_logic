from flask import Flask, request, jsonify;
from flask import render_template;
import serial;
import numpy as np
import skfuzzy as fuzz
import ratio_code as ratio
# import requests
# import schedule
# import time
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
wb = Workbook()
ws = wb.active
ws.append(['Temperature', 'Voltage1', 'Voltage2', 'Voltage3', 'Voltage4', 'Current1', 'Current2', 'Current3', 'Current4'])

# Constants
ser = serial.Serial('COM3', 9600, timeout=1)
#voltage_est = 18
#current_est = 3


@app.route("/")
def home():
    return render_template('login.html')

# define valid username and password
valid_username = 'admin'
valid_password = 'admin'

# Login api
@app.route('/login', methods=['POST'])
def login():
    # retrieve username and password from request form
    username = request.form.get('username')
    password = request.form.get('password')
    
    # check if username and password are valid
    if username == valid_username and password == valid_password:
        return jsonify({'success': True})
    else:
        return jsonify({'success': False})

@app.route("/arduinodata")
def data():
    return render_template('index.html')

#localhost:5000/data
@app.route('/data')
def get_data():
    try:
        # Read values from serial port
        line = ser.readline().decode().strip()
        temperature = 0
        voltage = 0
        oc_voltage=0
        current = 0
        vmpp = 16 #constant
        vcof = -0.0038 #constant
        
        if line != '':
            splitted = line.split(',')
            temperature = float(splitted[0])
            voltage = float(splitted[1])
            oc_voltage = float(splitted[2])
            current = float(splitted[3])
        
        # Write the temperature, voltage and current values to the excel file
        max_row = ws.max_row
        ws.append([temperature] + voltage + current)
        wb.save('data.xlsx')
                      
        # Return the temperature, voltage and current values as JSON
        # return jsonify({"temperature": temperature, "voltages": voltages, "currents": currents})
        print('Temperature:', temperature, 'Voltages:', voltage, 'Currents:', current)

        
        current = current
        # get irridiance value from the api.
        current_est = (119/1000)*1.36 #formula to find estimated current
        voltage = voltage
        voltage_est = vmpp*(1+(vcof*(temperature-25))) #formula to find estimated voltage
        oc_est = 21*(1+(vcof*(temperature-25))) #estimated open ckt voltage
        oc = oc_voltage
        ir, vr, ocr = ratio.calculate_ratios(current, current_est,voltage, voltage_est, oc, oc_est)
        print("ir:", ir)
        print("vr:", vr)
        print("ocr:", ocr)
        
        #range for the ratios
        x_vr = np.arange(0.95, 2.2, 0.01)
        x_ir = np.arange(0.95, 2.2, 0.01)
        x_ocr = np.arange(0.95, 2.2, 0.01)
        x_vio = np.arange(0.95, 2.143, 0.01)
        #print("x_vr: " + str(x_vr))
        #print("x_ir: " + str(x_ir))
        #print("x_ocr: " + str(x_ocr))
        # Generate fuzzy membership functions
        vrf0 = fuzz.trimf(x_vr, [0.99, 1.03, 1.06])
        vrf1 = fuzz.trimf(x_vr, [0.96, 1, 1.04 ])
        vrf2 = fuzz.trimf(x_vr, [1.02, 1.05, 1.08])
        vrf3 = fuzz.trimf(x_vr, [1.15, 1.215, 1.24])
        vrf4 = fuzz.trimf(x_vr, [1.41, 1.455, 1.50])
        vrf5 = fuzz.trimf(x_vr, [1.16, 1.23, 1.25])
        vrf6 = fuzz.trimf(x_vr, [1.50, 1.53, 1.57])
        vrf7 = fuzz.trimf(x_vr, [1.05, 1.175, 1.3])
        irf0 = fuzz.trimf(x_ir, [0.99, 1.03, 1.06])
        irf1 = fuzz.trimf(x_ir, [1.96, 1.99, 2.12])
        irf2 = fuzz.trimf(x_ir, [2.02, 2.08, 2.15])
        irf3 = fuzz.trimf(x_ir, [0.99, 1.000099, 1.0115])
        irf4 = fuzz.trimf(x_ir, [0.99, 1, 1.02])
        irf5 = fuzz.trimf(x_ir, [1.008485, 1.02000005, 1.04])
        irf6 = fuzz.trimf(x_ir, [1, 1.03, 1.05])
        irf7 = fuzz.trimf(x_ir, [0.98, 1.01, 1.04])
        ocrf0 = fuzz.trimf(x_ocr, [0.99, 1.03, 1.06])
        ocrf1 = fuzz.trimf(x_ocr, [0.99, 1.02, 1.04])
        ocrf2 = fuzz.trimf(x_ocr, [1.03, 1.06, 1.10])
        ocrf3 = fuzz.trimf(x_ocr, [1.16, 1.2, 1.21])
        ocrf4 = fuzz.trimf(x_ocr, [1.43, 1.46, 1.49])
        ocrf5 = fuzz.trimf(x_ocr, [1.17, 1.23, 1.26])
        ocrf6 = fuzz.trimf(x_ocr, [1.47, 1.52, 1.57])
        ocrf7 = fuzz.trimf(x_ocr, [1.06, 1.10, 1.15])
        no = fuzz.trimf(x_vio, [0.95, 1, 1.143])
        ps = fuzz.trimf(x_vio, [1, 1.143, 1.286 ])
        ps_bp = fuzz.trimf(x_vio, [1.143, 1.286, 1.429])
        tsc = fuzz.trimf(x_vio, [1.286, 1.429, 1.571])
        fsc = fuzz.trimf(x_vio, [1.429, 1.571, 1.714])
        toc = fuzz.trimf(x_vio, [1.571, 1.714, 1.857])
        foc = fuzz.trimf(x_vio, [1.714, 1.857,2])
        tld = fuzz.trimf(x_vio, [1.857,2, 2.143])
        # We need the activation of our fuzzy membership functions at these values.
        # This is what fuzz.interp_membership exists for!
        vr_level_f0 = fuzz.interp_membership(x_vr, vrf0, vr)
        vr_level_f1 = fuzz.interp_membership(x_vr, vrf1, vr)
        vr_level_f2 = fuzz.interp_membership(x_vr, vrf2, vr)
        vr_level_f3 = fuzz.interp_membership(x_vr, vrf3, vr)
        vr_level_f4 = fuzz.interp_membership(x_vr, vrf4, vr)
        vr_level_f5 = fuzz.interp_membership(x_vr, vrf5, vr)
        vr_level_f6 = fuzz.interp_membership(x_vr, vrf6, vr)
        vr_level_f7 = fuzz.interp_membership(x_vr, vrf7, vr)
        ir_level_f0 = fuzz.interp_membership(x_ir, irf0, ir)
        ir_level_f1 = fuzz.interp_membership(x_ir, irf1, ir)
        ir_level_f2 = fuzz.interp_membership(x_ir, irf2, ir)
        ir_level_f3 = fuzz.interp_membership(x_ir, irf3, ir)
        ir_level_f4 = fuzz.interp_membership(x_ir, irf4, ir)
        ir_level_f5 = fuzz.interp_membership(x_ir, irf5, ir)
        ir_level_f6 = fuzz.interp_membership(x_ir, irf6, ir)
        ir_level_f7 = fuzz.interp_membership(x_ir, irf7, ir)
        ocr_level_f0 = fuzz.interp_membership(x_ocr, ocrf0, ocr)
        ocr_level_f1 = fuzz.interp_membership(x_ocr, ocrf1, ocr)
        ocr_level_f2 = fuzz.interp_membership(x_ocr, ocrf2, ocr)
        ocr_level_f3 = fuzz.interp_membership(x_ocr, ocrf3, ocr)
        ocr_level_f4 = fuzz.interp_membership(x_ocr, ocrf4, ocr)
        ocr_level_f5 = fuzz.interp_membership(x_ocr, ocrf5, ocr)
        ocr_level_f6 = fuzz.interp_membership(x_ocr, ocrf6, ocr)
        ocr_level_f7 = fuzz.interp_membership(x_ocr, ocrf7, ocr)
        results = dict()
        # Now we take our rules and apply them.
        # Rule 1: vrf0 & irf0 & ocrf0 -> NO
        VI_combined = np.fmin(vr_level_f0, ir_level_f0)
        active_rule1 = np.fmin(ocr_level_f0, VI_combined )
        sys_activation_no = np.fmin(active_rule1, no)
        results['active_rule1'] = active_rule1
        # Rule 2: vrf1 & irf1 & ocrf1 -> PS
        VI_combined = np.fmin(vr_level_f1, ir_level_f1)
        active_rule2 = np.fmin(ocr_level_f1, VI_combined )
        sys_activation_ps = np.fmin(active_rule2, ps)
        results['active_rule2'] = active_rule2
        # Rule 3: vrf2 & irf2 & ocrf2 -> PS_BP
        VI_combined = np.fmin(vr_level_f2, ir_level_f2)
        active_rule3 = np.fmin(ocr_level_f2, VI_combined )
        sys_activation_ps_bp = np.fmin(active_rule2, ps_bp)
        results['active_rule3'] = active_rule3
        # Rule 4: vrf3 & irf3 & ocrf3 -> TSC
        VI_combined = np.fmin(vr_level_f3, ir_level_f3)
        active_rule4 = np.fmin(ocr_level_f3, VI_combined )
        sys_activation_tsc = np.fmin(active_rule4, tsc)
        results['active_rule4'] = active_rule4
        # Rule 5: vrf4 & irf4 & ocrf4 -> FSC
        VI_combined = np.fmin(vr_level_f4, ir_level_f4)
        active_rule5 = np.fmin(ocr_level_f4, VI_combined )
        sys_activation_fsc = np.fmin(active_rule5, fsc)
        results['active_rule5'] = active_rule5
        # Rule 6: vr5 & irf5 & ocrf5 -> TOC
        VI_combined = np.fmin(vr_level_f5, ir_level_f5)
        active_rule6 = np.fmin(ocr_level_f5, VI_combined )
        sys_activation_toc = np.fmin(active_rule6, toc)
        results['active_rule6'] = active_rule6
        # Rule 7: vrf6 & irf6 & ocrf6 -> FOC
        VI_combined = np.fmin(vr_level_f6, ir_level_f6)
        active_rule7 = np.fmin(ocr_level_f6, VI_combined )
        sys_activation_foc = np.fmin(active_rule2, foc)
        results['active_rule7'] = active_rule7
        # Rule 8: vrf7 & irf7 & ocrf7 -> TLD
        VI_combined = np.fmin(vr_level_f7, ir_level_f7)
        active_rule8 = np.fmin(ocr_level_f7, VI_combined )
        sys_activation_tld = np.fmin(active_rule8, tld)
        results['active_rule8'] = active_rule8
        sys0 = np.zeros_like(x_vio)
        # Aggregate all five output membership functions together
        aggregated = np.fmax(sys_activation_tld, np.fmax(sys_activation_foc, np.fmax(sys_activation_toc, np.fmax(sys_activation_fsc, np.fmax(sys_activation_tsc, np.fmax(sys_activation_ps_bp, np.fmax(sys_activation_ps, sys_activation_no)))))))
        #print("\naggregated: " + str(aggregated))
        # Calculate defuzzified result
        sys_stat = fuzz.defuzz(x_vio, aggregated, 'centroid')
        print("\nsys_stat: " + str(sys_stat))
        print(results)
        rules = {
            active_rule1 : 'No_fault',
            active_rule2 : 'Partial_shading',
            active_rule3 : 'PS_BP',
            active_rule4 : 'T_Short_ckt',
            active_rule5 : 'F_Short_ckt',
            active_rule6 : 't_open_ckt',
            active_rule7 : 'F_Open_ckt',
            active_rule8 : 'Leaves_dropping'
        }
        max_rule = max(active_rule1, active_rule2, active_rule3, active_rule4, active_rule5, active_rule6, active_rule7, active_rule8)
        if max_rule > 0:
            print("The active condition is:", rules[max_rule])
            return jsonify({"The active condition is:", rules[max_rule]})
        else:
            print("No active condition")
            return jsonify({"msg":"No active condition"})

    except KeyboardInterrupt:
        # Close the serial port on interrupt
        ser.close()

# def hit_api():
#     response = requests.get('http://localhost:5000/data')
#     print(response.json())

# schedule.every(5).minutes.do(hit_api)

# while True:
#     schedule.run_pending()
#     time.sleep(1)
    

if __name__ == '__main__':
    app.run(host='localhost', port=5000)
